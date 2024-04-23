import os, openpyxl, warnings, json

import openpyxl.workbook
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

"""
Script to update localization files
execute only in StreamingAssets folder!

"""
# check if right folder to run in
if os.getcwd().split('\\')[-1] != 'StreamingAssets':
    print('blyat')
    exit()

COLUMN_RU = 8
with open("language_dump.json", encoding='utf8') as f:
    language_dump = json.load(f)
with open('language_dump_backup.json', 'w', encoding='utf-8') as f:
    json.dump(language_dump, f, ensure_ascii=False)

# flag for updating
newStringsAdded = False
# parse folder and subfolders
for folder in os.walk(os.path.join(os.getcwd())):
    # filer non xlsx files
    tables = [x for x in folder[2] if x.startswith('Localization_')]
    path = folder[0]
    # for each file
    for t in tables:
        if t in language_dump['localization_file_list']['ignore']:
            continue
        print('checking',t)
        fileWasChanged = False
        # load it
        fullpath = path + "\\" + t
        wb = openpyxl.load_workbook(fullpath)
        ws = wb.worksheets[0]
        # check if file is new
        if t not in language_dump['localization_file_list']['files']:
            newStringsAdded = True
            print('found new file: ', t)
            language_dump['localization_file_list']['files'].append(t)
            ws.cell(row=1,column=COLUMN_RU).value = 'Русский'
        # check if file was changed
        if ws.cell(row=1,column=COLUMN_RU).value != 'Русский':
            ws.cell(row=1,column=COLUMN_RU).value = 'Русский'
        # iterate over rows starting with second one
        row = 2
        for key, cn, en, tcn, jp, kr, pt, ru, *values in ws.iter_rows(min_row=2, values_only=True):
            # check if not empty string
            if key == None:
                continue
            if en == None:
                en = ''
            if ru == None:
                ru = ''
            # check if it's new string
            if key not in language_dump:
                newStringsAdded = True
                print('new string: ', t, key)
                language_dump[key] = {'files':[t],'en':en,'ru':''}
            if language_dump[key]['en'] == None:
                language_dump[key]['en'] = ''
            if language_dump[key]['ru'] == None:
                language_dump[key]['ru'] = ''
            # check if same key suddenly not appeared elsewhere
            if t not in language_dump[key]['files']:
                newStringsAdded = True
                print('string appeared in another file: ', t, key)
                language_dump[key]['files'].append(t)
            # check if I changed something
            if ru != language_dump[key]['ru'] and language_dump[key]['ru'] != '':
                fileWasChanged = True
                translationChanged = True
                print('found change for', key + ':', f"\"{ru.replace('\n',' ')}\"", '=>', f"\"{language_dump[key]['ru'].replace('\n',' ')}\"")
                # update
                ws.cell(row, COLUMN_RU).value = language_dump[key]['ru']
            row += 1
            # check if there was change in en localization
            #continue
            if en != language_dump[key]['en']:
                # bug: files may have different values of one key, but so far seems that translation is consistent
                # doubles found:
                if key in ('易爆', '造成伤害', '1c6757aa97eb2a367697479776b9c18c', 'b3e80b2a4ec3fec39b9da30eaa3bc7b2','c44fd5492e9c0dfee96d7750597aaea9','44d61c38b0812f79231192d9417056e1'):
                    continue
                newStringsAdded = True
                print('found change for', key + ':', f"\"{language_dump[key]['en'].replace('\n',' ')}\""), '=>', f"\"{en.replace('\n',' ')}\""
                language_dump[key]['en'] = en
        # save changes
        if fileWasChanged:
            wb.save(fullpath)
            print('saved changes found')
        else:
            print('no changes found')
        wb.close()
print('done')

if newStringsAdded:
    # update dump if there was anything new
    with open('language_dump.json', 'w', encoding='utf-8') as f:
        json.dump(language_dump, f, ensure_ascii=False)